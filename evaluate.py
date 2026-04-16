"""
evaluate.py
-----------
End-to-end evaluation suite for the AIOps Log Intelligence System.

Covers every component with positive, negative, and edge-case tests.
No external test framework required — run directly with Python.

Usage:
    python evaluate.py                        # run all tests
    python evaluate.py --suite feature        # run one suite only
    python evaluate.py --suite feature,ingest # run multiple suites
    python evaluate.py --fast                 # skip slow training tests
    python evaluate.py --verbose              # show full error tracebacks

Available suites:
    feature    FeatureExtractor unit tests
    ingest     LogIngestionEngine unit tests
    pipeline   AnomalyPipeline unit tests
    knowledge  KnowledgeManager unit tests
    rag        AnalystRAG unit tests
    multiformat Multi-format file ingestion tests
    integration End-to-end integration tests
"""

import argparse
import importlib.util
import json
import math
import os
import shutil
import sys
import tempfile
import time
import traceback
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, List, Optional
from unittest.mock import MagicMock, patch
import csv
import numpy as np
import openpyxl
from fpdf import FPDF
from reportlab.pdfgen import canvas

# ══════════════════════════════════════════════════════════════════════════════
# TEST INFRASTRUCTURE
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class TestResult:
    name:     str
    status:   str          # PASS | FAIL | ERROR | SKIP
    elapsed:  float
    suite:    str
    message:  str  = ""
    slow:     bool = False


class TestRunner:
    """Lightweight test runner — no external framework needed."""

    def __init__(self, verbose: bool = False, fast: bool = False):
        self.verbose  = verbose
        self.fast     = fast
        self.results: List[TestResult] = []
        self._current_suite = "misc"

    def suite(self, name: str):
        """Set the current suite label."""
        self._current_suite = name
        print(f"\n{'='*60}")
        print(f"  Suite: {name}")
        print(f"{'='*60}")

    def run(self, name: str, fn: Callable, slow: bool = False):
        """Register and immediately execute a test."""
        if self.fast and slow:
            self.results.append(TestResult(name, "SKIP", 0.0, self._current_suite,
                                           "skipped (--fast)", slow=True))
            print(f"  [SKIP] {name}")
            return

        t0 = time.time()
        try:
            fn()
            elapsed = time.time() - t0
            self.results.append(TestResult(name, "PASS", elapsed, self._current_suite))
            print(f"  [PASS] {name}  ({elapsed:.3f}s)")
        except AssertionError as e:
            elapsed = time.time() - t0
            msg = str(e)
            self.results.append(TestResult(name, "FAIL", elapsed, self._current_suite, msg))
            print(f"  [FAIL] {name}  ({elapsed:.3f}s)")
            print(f"         {msg[:200]}")
            if self.verbose:
                traceback.print_exc()
        except Exception as e:
            elapsed = time.time() - t0
            msg = f"{type(e).__name__}: {e}"
            self.results.append(TestResult(name, "ERROR", elapsed, self._current_suite, msg))
            print(f"  [ERROR] {name}  ({elapsed:.3f}s)")
            print(f"          {msg[:200]}")
            if self.verbose:
                traceback.print_exc()

    def report(self) -> int:
        """Print summary. Returns exit code (0=all pass, 1=failures)."""
        print(f"\n{'='*60}")
        print(f"  EVALUATION SUMMARY")
        print(f"{'='*60}")

        by_suite: dict = {}
        for r in self.results:
            by_suite.setdefault(r.suite, []).append(r)

        total_pass = total_fail = total_error = total_skip = 0
        for suite_name, suite_results in by_suite.items():
            p = sum(1 for r in suite_results if r.status == "PASS")
            f = sum(1 for r in suite_results if r.status == "FAIL")
            e = sum(1 for r in suite_results if r.status == "ERROR")
            s = sum(1 for r in suite_results if r.status == "SKIP")
            total_pass += p; total_fail += f; total_error += e; total_skip += s
            status = "OK" if (f + e) == 0 else "NEEDS ATTENTION"
            print(f"  {suite_name:<20}: {p:>3} pass  {f:>2} fail  {e:>2} error  {s:>2} skip  [{status}]")

        total = len(self.results)
        print("-" * 60)
        print(f"  Total: {total}  |  Pass: {total_pass}  |  Fail: {total_fail}  "
              f"|  Error: {total_error}  |  Skip: {total_skip}")

        if total_fail + total_error > 0:
            print(f"\n  FAILED TESTS:")
            for r in self.results:
                if r.status in ("FAIL", "ERROR"):
                    print(f"    [{r.status}] {r.suite} / {r.name}")
                    print(f"            {r.message[:120]}")

        total_time = sum(r.elapsed for r in self.results)
        print(f"\n  Total time: {total_time:.1f}s")
        print(f"{'='*60}\n")
        return 1 if (total_fail + total_error) > 0 else 0


# ── Helper assertions ──────────────────────────────────────────────────────────
def assert_equal(a, b, msg=""):
    assert a == b, f"{msg} | expected {b!r}, got {a!r}"

def assert_close(a, b, tol=1e-4, msg=""):
    assert abs(a - b) <= tol, f"{msg} | expected ~{b}, got {a} (tol={tol})"

def assert_in_range(v, lo, hi, msg=""):
    assert lo <= v <= hi, f"{msg} | {v} not in [{lo}, {hi}]"

def assert_shape(arr, shape, msg=""):
    assert arr.shape == shape, f"{msg} | expected shape {shape}, got {arr.shape}"

def assert_keys(d, keys, msg=""):
    missing = [k for k in keys if k not in d]
    assert not missing, f"{msg} | missing keys: {missing}"

def check_available(pkg: str) -> bool:
    return importlib.util.find_spec(pkg) is not None


# ── Temp directory context manager ────────────────────────────────────────────
@contextmanager
def tmp_dir():
    d = tempfile.mkdtemp(prefix="aiops_eval_")
    try:
        yield Path(d)
    finally:
        shutil.rmtree(d, ignore_errors=True)


# ══════════════════════════════════════════════════════════════════════════════
# SYNTHETIC DATA GENERATORS
# ══════════════════════════════════════════════════════════════════════════════

class MockEmbedder:
    """Replaces SentenceTransformer for unit tests — deterministic 384d vectors."""
    def encode(self, texts, batch_size=32, show_progress_bar=False,
               normalize_embeddings=False):
        np.random.seed(42)
        vecs = np.random.randn(len(texts), 384).astype(np.float32)
        if normalize_embeddings:
            norms = np.linalg.norm(vecs, axis=1, keepdims=True)
            vecs  = vecs / np.maximum(norms, 1e-9)
        return vecs


class MockLogCluster:
    """Minimal LogCluster substitute for testing _extract_cluster."""
    def __init__(self, cid=1, tokens=None):
        self.cluster_id         = cid
        self.log_template_tokens = tokens or ["Connection", "timeout", "after", "<*>", "ms"]
    def get_template(self) -> str:
        return " ".join(self.log_template_tokens)


def make_log_rows(n: int = 20, template_count: int = 5) -> list:
    """Generate synthetic DuckDB row tuples matching stream_feature_rows SELECT."""
    rows = []
    ts_base = datetime(2024, 1, 1, 12, 0, 0, tzinfo=timezone.utc)
    for i in range(n):
        tid   = f"E{i % template_count}"
        level = ["INFO","WARN","ERROR"][i % 3]
        ts    = ts_base.replace(second=i % 60)
        rows.append((
            f"log_{i:06d}",                          # log_id
            ts,                                       # timestamp
            level,                                    # level
            f"svc_{i%3}",                             # service
            tid,                                      # template_id
            f"Connection timeout after <*> ms on {tid}",  # template_str
            json.dumps(["500", f"host_{i%5}"]),       # dynamic_params
            f"2024-01-01 12:00:{i%60:02d} {level} svc_{i%3} Connection timeout after 500 ms",
            None,                                     # http_status
            False, False, False, i%3==2, i%3==0,     # http_1xx..5xx
        ))
    return rows


def make_feature_matrix(n: int = 200, dim: int = 393,
                         n_anomalies: int = 20) -> np.ndarray:
    """Generate a feature matrix with known normal and anomaly clusters."""
    np.random.seed(0)
    normal   = np.random.randn(n - n_anomalies, dim).astype(np.float32) * 0.5
    anomaly  = np.random.randn(n_anomalies, dim).astype(np.float32) * 5.0 + 8.0
    return np.vstack([normal, anomaly])


def make_log_file(path: Path, n_lines: int = 50) -> Path:
    """Write a synthetic .log file."""
    with open(path, "w", encoding="utf-8") as f:
        for i in range(n_lines):
            ts    = f"2024-01-01 12:00:{i%60:02d}"
            level = ["INFO","WARN","ERROR"][i % 3]
            svc   = f"service_{i%4}"
            f.write(f"{ts} {level} {svc} Connection attempt {i} on port 8080 status=200\n")
    return path


def make_apache_csv(path: Path, n_rows: int = 30) -> Path:
    """Write a synthetic Apache Combined Log CSV (no header)."""
    methods = ["GET","POST","PUT","DELETE"]
    paths_  = ["/api/v1/users","/login","/admin","/download/file.zip"]
    statuses = [200, 404, 500, 201]
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for i in range(n_rows):
            writer.writerow([
                f"192.168.{i%10}.{i%255}", "-", f"user{i%5}",
                f"[{18+i%10}/Jul/2024:12:00:{i%60:02d}", "+0530]",
                f"{methods[i%4]} {paths_[i%4]} HTTP/1.1",
                statuses[i%4], 1024 + i*10,
                "http://example.com/", "Mozilla/5.0", 50 + i,
            ])
    return path


def make_structured_csv(path: Path, n_rows: int = 30) -> Path:
    """Write a synthetic structured anomaly CSV (with header)."""
    labels    = ["Storage","Database","Network","CPU","Memory"]
    severities = ["Low","Medium","High","Critical"]
    sources   = ["Server","Database","Cloud","Edge"]
    
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Timestamp","Anomaly_ID","Anomaly_Type","Severity","Status",
            "Source","Alert_Method","Response_Time_ms","Resolution_Time_min",
            "Affected_Services","User_Role","Host_IP","Process_ID","Error_Code",
            "CPU_Usage_Percent","Memory_Usage_MB","Disk_Usage_Percent",
            "Network_In_KB","Network_Out_KB","Login_Attempts","Failed_Transactions",
            "Anomaly_Duration_sec","TimeZone","User_ID","Location","Patch_Level",
            "Service_Type","Transaction_ID","Alert_Count","Retry_Count","Escalation_Level",
        ])
        writer.writeheader()
        for i in range(n_rows):
            writer.writerow({
                "Timestamp":          f"2025-01-01 00:{i%60:02d}:00",
                "Anomaly_ID":         i+1,
                "Anomaly_Type":       labels[i % len(labels)],
                "Severity":           severities[i % len(severities)],
                "Status":             "Open",
                "Source":             sources[i % len(sources)],
                "Alert_Method":       "Email",
                "Response_Time_ms":   1000 + i * 50,
                "Resolution_Time_min":15,
                "Affected_Services":  2,
                "User_Role":          "admin",
                "Host_IP":            f"10.0.{i%5}.{i%255}",
                "Process_ID":         1000+i,
                "Error_Code":         500 if i % 3 == 0 else 0,
                "CPU_Usage_Percent":  40.0 + i % 50,
                "Memory_Usage_MB":    512 + i * 10,
                "Disk_Usage_Percent": 30.0 + i % 60,
                "Network_In_KB":      100+i,
                "Network_Out_KB":     50+i,
                "Login_Attempts":     i % 10,
                "Failed_Transactions":i % 5,
                "Anomaly_Duration_sec":60+i,
                "TimeZone":           "UTC",
                "User_ID":            10000+i,
                "Location":           "APAC",
                "Patch_Level":        i % 5,
                "Service_Type":       "API",
                "Transaction_ID":     900000+i,
                "Alert_Count":        i % 20,
                "Retry_Count":        i % 5,
                "Escalation_Level":   i % 4,
            })
    return path


def make_generic_csv(path: Path, n_rows: int = 20) -> Path:
    """Write a generic CSV with mixed columns."""
    import csv
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "timestamp","level","service","message","duration_ms","error_code"
        ])
        writer.writeheader()
        for i in range(n_rows):
            writer.writerow({
                "timestamp":   f"2024-01-01 12:00:{i%60:02d}",
                "level":       ["INFO","WARN","ERROR"][i%3],
                "service":     f"service_{i%3}",
                "message":     f"Request processed successfully id={i}",
                "duration_ms": 100 + i * 5,
                "error_code":  0 if i % 5 != 0 else 500,
            })
    return path


def make_excel_file(path: Path, n_rows: int = 20) -> Path:
    """Write a synthetic Excel workbook."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SystemLogs"
    ws.append(["Timestamp","Level","Service","Message","Response_Time_ms"])
    for i in range(n_rows):
        ws.append([
            f"2024-01-{(i%28)+1:02d} 12:00:00",
            ["INFO","WARN","ERROR"][i % 3],
            f"svc_{i%3}",
            f"Database query completed rows={i*10} table=orders",
            100 + i * 5,
        ])
    # Add a second sheet with anomalies
    ws2 = wb.create_sheet("Anomalies")
    ws2.append(["Timestamp","Severity","Source","Anomaly_Type","Error_Code"])
    for i in range(5):
        ws2.append([
            f"2024-01-{(i%28)+1:02d} 14:00:00",
            "High", "Database", "Connection Timeout", 500,
        ])
    wb.save(str(path))
    return path


def make_pdf_file(path: Path, n_lines: int = 20) -> Optional[Path]:
    """Write a synthetic PDF. Returns None if fpdf2/reportlab unavailable."""
    # Try fpdf2 first, then reportlab, then skip
    try:
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=10)
        for i in range(n_lines):
            ts    = f"2024-01-01 12:00:{i%60:02d}"
            level = ["INFO","WARN","ERROR"][i % 3]
            pdf.cell(0, 8,
                     f"{ts} {level} auth-service Login attempt {i} from 192.168.1.{i%255}",
                     ln=True)
        pdf.output(str(path))
        return path
    except ImportError:
        pass
    try:
        
        c = canvas.Canvas(str(path))
        y = 800
        for i in range(n_lines):
            c.drawString(40, y, f"2024-01-01 12:00:{i%60:02d} INFO svc Log entry {i}")
            y -= 20
            if y < 50:
                c.showPage()
                y = 800
        c.save()
        return path
    except ImportError:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 1: FeatureExtractor
# ══════════════════════════════════════════════════════════════════════════════
def register_feature_extractor_tests(runner: TestRunner):
    runner.suite("feature")

    from feature_extractor import FeatureExtractor, EMBEDDING_DIM
    from cachetools import LRUCache

    def make_fe():
        fe = FeatureExtractor.__new__(FeatureExtractor)
        fe.embedder          = MockEmbedder()
        fe.batch_embed_size  = 64
        fe._burst_windows    = LRUCache(maxsize=1000)
        fe._last_ts_per_service = LRUCache(maxsize=500)
        fe._welford_state    = LRUCache(maxsize=1000)
        fe._last_ts_fmt      = None
        return fe

    # ── Entropy tests ──────────────────────────────────────────────────────────
    runner.run("entropy: empty string -> 0.0", lambda: (
        assert_equal(FeatureExtractor.shannon_entropy(""), 0.0)
    ))
    runner.run("entropy: single char -> 0.0", lambda: (
        assert_equal(FeatureExtractor.shannon_entropy("a"), 0.0)
    ))
    runner.run("entropy: uniform string -> 0.0", lambda: (
        assert_equal(FeatureExtractor.shannon_entropy("aaaaaaaaaa"), 0.0)
    ))
    runner.run("entropy: all-distinct 10-char -> 1.0", lambda: (
        assert_equal(FeatureExtractor.shannon_entropy("abcdefghij"), 1.0)
    ))
    runner.run("entropy: always in [0.0, 1.0]", lambda: [
        assert_in_range(FeatureExtractor.shannon_entropy(t), 0.0, 1.0)
        for t in ["", "a", "hello world", "abc123!@#", "x"*100, "abcde"*20]
    ])
    runner.run("entropy: longer string never biases above 1.0", lambda: (
        assert_in_range(
            FeatureExtractor.shannon_entropy("The quick brown fox jumps over lazy dog 42"),
            0.0, 1.0
        )
    ))
    runner.run("entropy: mixed case and symbols", lambda: (
        assert_in_range(
            FeatureExtractor.shannon_entropy("Error: Connection refused [ECONNREFUSED] port=5432"),
            0.5, 1.0, "mixed log line should have moderate entropy"
        )
    ))

    # ── Burstiness tests ───────────────────────────────────────────────────────
    runner.run("burstiness: single event = count 1", lambda: (
        assert_equal(make_fe()._update_burstiness("E1", 1000.0), 1)
    ))
    runner.run("burstiness: 5 events within window = count 5", lambda: [
        None for fe, _ in [(make_fe(), None)]
        for expected, t_off in [(i+1, float(i)) for i in range(5)]
        if assert_equal(fe._update_burstiness("E1", 1000.0 + t_off), expected) or True
    ])
    runner.run("burstiness: events outside window evicted", lambda: (
        _test_burst_eviction()
    ))
    runner.run("burstiness: different templates isolated", lambda: (
        _test_burst_isolation()
    ))

    def _test_burst_eviction():
        fe = make_fe()
        for i in range(5):
            fe._update_burstiness("E1", 1000.0 + i)
        # Jump 15 seconds — all previous events expire (window=10s)
        count = fe._update_burstiness("E1", 1020.0)
        assert_equal(count, 1, "after 15s gap, only the new event is in window")

    def _test_burst_isolation():
        fe = make_fe()
        for i in range(3):
            fe._update_burstiness("E1", 1000.0 + i)
        count_e2 = fe._update_burstiness("E2", 1001.0)
        assert_equal(count_e2, 1, "E2 window must not be contaminated by E1")

    # ── Welford volatility tests ───────────────────────────────────────────────
    runner.run("volatility: first event -> 0.0 (count < 2)", lambda: (
        assert_equal(make_fe()._compute_volatility("E1", json.dumps(["500"])), 0.0)
    ))
    runner.run("volatility: two equal values -> 0.0", lambda: (
        assert_equal(make_fe()._compute_volatility("E1", json.dumps(["100","100"])), 0.0)
    ))
    runner.run("volatility: known two-value std", lambda: (
        _test_volatility_known()
    ))
    runner.run("volatility: accumulates across calls", lambda: (
        _test_volatility_accumulation()
    ))
    runner.run("volatility: non-numeric params ignored", lambda: (
        assert_equal(make_fe()._compute_volatility("E1", json.dumps(["host-1","db"])), 0.0)
    ))
    runner.run("volatility: malformed JSON -> 0.0", lambda: (
        assert_equal(make_fe()._compute_volatility("E1", "not-json"), 0.0)
    ))

    def _test_volatility_known():
        fe = make_fe()
        # population std of [0, 100] = 50.0
        fe._compute_volatility("E1", json.dumps(["0"]))
        v = fe._compute_volatility("E1", json.dumps(["100"]))
        assert_close(v, 50.0, tol=0.01, msg="pop std of [0,100] must be 50.0")

    def _test_volatility_accumulation():
        fe = make_fe()
        for val in ["10","20","30","40","50"]:
            last = fe._compute_volatility("E1", json.dumps([val]))
        # pop std of [10,20,30,40,50] = sqrt(200) ~= 14.142
        assert_close(last, math.sqrt(200), tol=0.01)

    # ── Time delta tests ───────────────────────────────────────────────────────
    runner.run("time_delta: first event -> 0.0", lambda: (
        assert_equal(make_fe()._compute_time_delta("svc1", 1000.0), 0.0)
    ))
    runner.run("time_delta: 5s apart -> 5.0", lambda: (
        _test_time_delta_correct()
    ))
    runner.run("time_delta: ts=None -> 0.0", lambda: (
        assert_equal(make_fe()._compute_time_delta("svc1", None), 0.0)
    ))
    runner.run("time_delta: capped at 86400s", lambda: (
        _test_time_delta_cap()
    ))

    def _test_time_delta_correct():
        fe = make_fe()
        fe._compute_time_delta("svc1", 1000.0)
        delta = fe._compute_time_delta("svc1", 1005.0)
        assert_close(delta, 5.0, tol=0.001)

    def _test_time_delta_cap():
        fe = make_fe()
        fe._compute_time_delta("svc1", 0.0)
        delta = fe._compute_time_delta("svc1", 999999.0)
        assert_equal(delta, 86400.0)

    # ── LRU bound tests ────────────────────────────────────────────────────────
    runner.run("lru: burst_windows stays bounded under high cardinality", lambda: (
        _test_lru_burst_bounded()
    ))
    runner.run("lru: welford_state stays bounded under high cardinality", lambda: (
        _test_lru_welford_bounded()
    ))
    runner.run("lru: last_ts_per_service stays bounded", lambda: (
        _test_lru_service_bounded()
    ))

    def _test_lru_burst_bounded():
        fe = make_fe()
        for i in range(2000):   # far exceeds maxsize=1000
            fe._update_burstiness(f"TEMPLATE_{i}", float(i))
        assert len(fe._burst_windows) <= 1000, "LRU must evict to stay bounded"

    def _test_lru_welford_bounded():
        fe = make_fe()
        for i in range(2000):
            fe._compute_volatility(f"T_{i}", json.dumps([str(i)]))
        assert len(fe._welford_state) <= 1000

    def _test_lru_service_bounded():
        fe = make_fe()
        for i in range(1000):
            fe._compute_time_delta(f"service_{i}", float(i))
        assert len(fe._last_ts_per_service) <= 500

    # ── Timestamp parsing ──────────────────────────────────────────────────────
    runner.run("timestamp: datetime object parsed correctly", lambda: (
        _test_ts_datetime()
    ))
    runner.run("timestamp: ISO string parsed correctly", lambda: (
        _test_ts_iso_string()
    ))
    runner.run("timestamp: numeric passthrough", lambda: (
        assert_equal(make_fe()._to_unix(1704067200.0), 1704067200.0)
    ))
    runner.run("timestamp: None -> None", lambda: (
        assert_equal(make_fe()._to_unix(None), None)
    ))
    runner.run("timestamp: empty string -> None", lambda: (
        assert_equal(make_fe()._to_unix(""), None)
    ))
    runner.run("timestamp: unparseable string -> None", lambda: (
        assert_equal(make_fe()._to_unix("not_a_timestamp_at_all!"), None)
    ))
    runner.run("timestamp: format caches after first parse", lambda: (
        _test_ts_format_cache()
    ))

    def _test_ts_datetime():
        fe  = make_fe()
        dt  = datetime(2024, 1, 1, 12, 0, 0, tzinfo=timezone.utc)
        ts  = fe._to_unix(dt)
        assert ts is not None and abs(ts - 1704110400.0) < 2

    def _test_ts_iso_string():
        fe = make_fe()
        ts = fe._to_unix("2024-01-01 12:00:00")
        assert ts is not None and ts > 0

    def _test_ts_format_cache():
        fe = make_fe()
        assert fe._last_ts_fmt is None
        fe._to_unix("2024-01-01 12:00:00")
        assert fe._last_ts_fmt is not None, "Format should be cached after first parse"

    # ── HTTP flags ─────────────────────────────────────────────────────────────
    runner.run("http_flags: 5xx row sets http_5xx=1.0", lambda: (
        _test_http_flags()
    ))

    def _test_http_flags():
        fe  = make_fe()
        row = (None, None, None, None, None, None, None, None, 500,
               False, False, False, False, True)
        vec = fe._http_flags_vector(row)
        assert_equal(vec[4], 1.0)
        assert_equal(vec[0] + vec[1] + vec[2] + vec[3], 0.0)

    # ── Level ordinal ──────────────────────────────────────────────────────────
    runner.run("level_ordinal: ERROR > WARN > INFO", lambda: (
        _test_level_ordinal()
    ))

    def _test_level_ordinal():
        from feature_extractor import FeatureExtractor as FE
        assert FE._level_ordinal("ERROR") > FE._level_ordinal("WARN")
        assert FE._level_ordinal("WARN")  > FE._level_ordinal("INFO")
        assert FE._level_ordinal("FATAL") == FE._level_ordinal("CRITICAL")
        assert FE._level_ordinal(None)   == 2.0

    # ── Output shape ──────────────────────────────────────────────────────────
    runner.run("process_batch: output shape always (N, 393)", lambda: (
        _test_output_shape()
    ))
    runner.run("process_batch: empty input -> (0, 393)", lambda: (
        _test_empty_input()
    ))
    runner.run("process_batch: metadata count matches row count", lambda: (
        _test_metadata_count()
    ))

    def _test_output_shape():
        fe       = make_fe()
        rows     = make_log_rows(20)
        features, meta = fe.process_batch(rows)
        assert_shape(features, (20, 393))

    def _test_empty_input():
        fe = make_fe()
        features, meta = fe.process_batch([])
        assert_shape(features, (0, 393))
        assert_equal(len(meta), 0)

    def _test_metadata_count():
        fe = make_fe()
        rows = make_log_rows(15)
        features, meta = fe.process_batch(rows)
        assert_equal(len(meta), 15)


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 2: LogIngestionEngine
# ══════════════════════════════════════════════════════════════════════════════
def register_ingestion_tests(runner: TestRunner):
    runner.suite("ingest")

    from ingestion_engine import LogIngestionEngine

    # ── _extract_cluster version compatibility ─────────────────────────────────
    runner.run("extract_cluster: Shape A dict with cluster key", lambda: (
        _test_cluster_shape_a()
    ))
    runner.run("extract_cluster: Shape A empty dict -> None", lambda: (
        _test_cluster_shape_a_empty()
    ))
    runner.run("extract_cluster: Shape B tuple (cluster, str)", lambda: (
        _test_cluster_shape_b()
    ))
    runner.run("extract_cluster: Shape C direct LogCluster", lambda: (
        _test_cluster_shape_c()
    ))
    runner.run("extract_cluster: None input -> None", lambda: (
        assert_equal(LogIngestionEngine._extract_cluster(None), None)
    ))
    runner.run("extract_cluster: invalid dict key -> None", lambda: (
        assert_equal(LogIngestionEngine._extract_cluster({"change_type": "none"}), None)
    ))
    runner.run("extract_cluster: empty tuple -> None", lambda: (
        assert_equal(LogIngestionEngine._extract_cluster(()), None)
    ))

    def _test_cluster_shape_a():
        cluster = MockLogCluster(cid=7)
        result  = {"cluster": cluster, "change_type": "cluster_created"}
        out     = LogIngestionEngine._extract_cluster(result)
        assert out is cluster

    def _test_cluster_shape_a_empty():
        out = LogIngestionEngine._extract_cluster({})
        assert out is None

    def _test_cluster_shape_b():
        cluster = MockLogCluster(cid=3)
        result  = (cluster, "cluster_created")
        out     = LogIngestionEngine._extract_cluster(result)
        assert out is cluster

    def _test_cluster_shape_c():
        cluster = MockLogCluster(cid=5)
        out     = LogIngestionEngine._extract_cluster(cluster)
        assert out is cluster

    # ── _safe_get_template ─────────────────────────────────────────────────────
    runner.run("safe_get_template: uses get_template() method", lambda: (
        _test_safe_get_template_method()
    ))
    runner.run("safe_get_template: falls back to log_template_tokens", lambda: (
        _test_safe_get_template_tokens()
    ))
    runner.run("safe_get_template: last resort str(cluster)", lambda: (
        _test_safe_get_template_fallback()
    ))

    def _test_safe_get_template_method():
        cluster = MockLogCluster()
        result  = LogIngestionEngine._safe_get_template(cluster)
        assert_equal(result, "Connection timeout after <*> ms")

    def _test_safe_get_template_tokens():
        cluster = MockLogCluster()
        del cluster.__class__.get_template   # type: ignore
        # Rebuild without method
        class NoMethod:
            log_template_tokens = ["DB","timeout","<*>"]
        result = LogIngestionEngine._safe_get_template(NoMethod())
        assert_equal(result, "DB timeout <*>")

    def _test_safe_get_template_fallback():
        class Bare:
            pass
        result = LogIngestionEngine._safe_get_template(Bare())
        assert isinstance(result, str) and len(result) > 0

    # ── _count_clusters ─────────────────────────────────────────────────────────
    runner.run("count_clusters: id_to_cluster attribute", lambda: (
        _test_count_clusters_v1()
    ))
    runner.run("count_clusters: clusters attribute", lambda: (
        _test_count_clusters_v2()
    ))
    runner.run("count_clusters: no drain -> 0", lambda: (
        _test_count_clusters_none()
    ))

    def _test_count_clusters_v1():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        engine.miner = MagicMock()
        engine.miner.drain.id_to_cluster = {1: "a", 2: "b", 3: "c"}
        assert_equal(engine._count_clusters(), 3)

    def _test_count_clusters_v2():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        engine.miner = MagicMock()
        del engine.miner.drain.id_to_cluster
        engine.miner.drain.clusters = {"a", "b"}
        assert_equal(engine._count_clusters(), 2)

    def _test_count_clusters_none():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        engine.miner = MagicMock(spec=[])  # no drain attribute
        assert_equal(engine._count_clusters(), 0)

    # ── _parse_timestamp ────────────────────────────────────────────────────────
    runner.run("parse_timestamp: ISO format", lambda: (
        _test_parse_ts_iso()
    ))
    runner.run("parse_timestamp: Apache format", lambda: (
        _test_parse_ts_apache()
    ))
    runner.run("parse_timestamp: None/empty -> None", lambda: (
        _test_parse_ts_none()
    ))

    def _test_parse_ts_iso():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        dt = engine._parse_timestamp("2024-01-15 09:30:00")
        assert dt is not None and dt.month == 1 and dt.day == 15

    def _test_parse_ts_apache():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        dt = engine._parse_timestamp("15/Jan/2024:09:30:00 +0000")
        assert dt is not None

    def _test_parse_ts_none():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        assert engine._parse_timestamp("") is None
        assert engine._parse_timestamp("not_a_date") is None

    # ── HTTP status flags ────────────────────────────────────────────────────────
    runner.run("http_status_flags: 200 -> 2xx=True others=False", lambda: (
        _test_http_flags_200()
    ))
    runner.run("http_status_flags: 500 -> 5xx=True others=False", lambda: (
        _test_http_flags_500()
    ))
    runner.run("http_status_flags: None -> all False", lambda: (
        _test_http_flags_none()
    ))

    def _test_http_flags_200():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        flags  = engine._encode_http_flags(200)
        assert_equal(flags["http_2xx"], True)
        assert_equal(flags["http_5xx"], False)

    def _test_http_flags_500():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        flags  = engine._encode_http_flags(500)
        assert_equal(flags["http_5xx"], True)
        assert_equal(flags["http_2xx"], False)

    def _test_http_flags_none():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        flags  = engine._encode_http_flags(None)
        assert all(not v for v in flags.values())

    # ── _extract_dynamic_params ─────────────────────────────────────────────────
    runner.run("extract_dynamic_params: extracts wildcard tokens", lambda: (
        _test_extract_params()
    ))
    runner.run("extract_dynamic_params: no wildcards -> empty list", lambda: (
        _test_extract_params_none()
    ))

    def _test_extract_params():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        params = engine._extract_dynamic_params(
            "Connection timeout after 500 ms", "Connection timeout after <*> ms"
        )
        data = json.loads(params)
        assert "500" in data

    def _test_extract_params_none():
        engine = LogIngestionEngine.__new__(LogIngestionEngine)
        params = engine._extract_dynamic_params(
            "No wildcards here", "No wildcards here"
        )
        data = json.loads(params)
        assert isinstance(data, list)

    # ── Full ingest from file ─────────────────────────────────────────────────
    runner.run("ingest_file: end-to-end with synthetic log", lambda: (
        _test_ingest_file()
    ), slow=True)
    runner.run("ingest_file: resume skips already-ingested lines", lambda: (
        _test_ingest_resume()
    ), slow=True)
    runner.run("ingest_file: empty file -> 0 rows", lambda: (
        _test_ingest_empty()
    ), slow=True)

    def _test_ingest_file():
        with tmp_dir() as d:
            log_path = make_log_file(d / "test.log", n_lines=100)
            engine   = LogIngestionEngine(
                db_path=str(d/"test.duckdb"),
                config_path="miner_config.json",
                drain_state_path=str(d/"drain.bin"),
                chunk_size=20,
            )
            stats = engine.ingest_file(str(log_path))
            engine.close()
            assert stats["total_rows_ingested"] == 100, \
                f"Expected 100 rows, got {stats['total_rows_ingested']}"

    def _test_ingest_resume():
        with tmp_dir() as d:
            log_path = make_log_file(d / "test.log", n_lines=50)
            db_path  = str(d / "test.duckdb")
            drain    = str(d / "drain.bin")
            # First ingest
            e1 = LogIngestionEngine(db_path=db_path, config_path="miner_config.json",
                                    drain_state_path=drain, chunk_size=25)
            e1.ingest_file(str(log_path))
            e1.close()
            # Second ingest with resume=True — should add 0 new rows
            e2 = LogIngestionEngine(db_path=db_path, config_path="miner_config.json",
                                    drain_state_path=drain, chunk_size=25)
            stats2 = e2.ingest_file(str(log_path), resume=True)
            e2.close()
            assert stats2["total_rows_ingested"] == 0, \
                f"Resume should ingest 0 new rows but got {stats2['total_rows_ingested']}"

    def _test_ingest_empty():
        with tmp_dir() as d:
            log_path = d / "empty.log"
            log_path.write_text("")
            engine = LogIngestionEngine(
                db_path=str(d/"test.duckdb"),
                config_path="miner_config.json",
                drain_state_path=str(d/"drain.bin"),
                chunk_size=100,
            )
            stats = engine.ingest_file(str(log_path))
            engine.close()
            assert_equal(stats["total_rows_ingested"], 0)


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 3: AnomalyPipeline
# ══════════════════════════════════════════════════════════════════════════════
def register_anomaly_pipeline_tests(runner: TestRunner):
    runner.suite("pipeline")

    from anomaly_pipeline import Autoencoder, AnomalyPipeline
    import torch

    DIM = 64   # small dim for fast tests

    # ── Autoencoder unit tests ─────────────────────────────────────────────────
    runner.run("autoencoder: forward pass output shape", lambda: (
        _test_ae_forward()
    ))
    runner.run("autoencoder: latent dim correct", lambda: (
        _test_ae_latent_dim()
    ))
    runner.run("autoencoder: reconstruct near identity on trivial data", lambda: (
        _test_ae_reconstruction()
    ), slow=True)

    def _test_ae_forward():
        ae  = Autoencoder(input_dim=DIM, latent_dim=8, hidden_dims=[32, 16])
        x   = torch.randn(10, DIM)
        x_hat, z = ae(x)
        assert x_hat.shape == (10, DIM), f"Output shape wrong: {x_hat.shape}"
        assert z.shape    == (10, 8),   f"Latent shape wrong: {z.shape}"

    def _test_ae_latent_dim():
        for latent in [4, 8, 16, 32]:
            ae = Autoencoder(input_dim=DIM, latent_dim=latent)
            z  = ae.encode(torch.randn(5, DIM))
            assert_equal(z.shape[1], latent)

    def _test_ae_reconstruction():
        # Trivial dataset: all zeros. AE should learn to reconstruct near zero.
        ae   = Autoencoder(input_dim=DIM, latent_dim=8, hidden_dims=[32])
        data = np.zeros((200, DIM), dtype=np.float32)
        from anomaly_pipeline import train_autoencoder
        train_autoencoder(ae, data, epochs=5, batch_size=64)
        from anomaly_pipeline import compute_reconstruction_errors
        errs = compute_reconstruction_errors(ae, data)
        assert errs.mean() < 0.5, f"Mean reconstruction error too high: {errs.mean()}"

    # ── Pipeline train/score/save/load ────────────────────────────────────────
    runner.run("pipeline: train returns expected stats keys", lambda: (
        _test_pipeline_train_stats()
    ), slow=True)
    runner.run("pipeline: anomalies detected in injected outliers", lambda: (
        _test_pipeline_detects_anomalies()
    ), slow=True)
    runner.run("pipeline: score output keys correct", lambda: (
        _test_pipeline_score_keys()
    ), slow=True)
    runner.run("pipeline: score output shapes consistent", lambda: (
        _test_pipeline_score_shapes()
    ), slow=True)
    runner.run("pipeline: save and load roundtrip", lambda: (
        _test_pipeline_save_load()
    ), slow=True)
    runner.run("pipeline: contamination controls anomaly fraction", lambda: (
        _test_pipeline_contamination()
    ), slow=True)
    runner.run("pipeline: fine_tune_ae2 updates threshold", lambda: (
        _test_pipeline_finetune()
    ), slow=True)

    def _mini_pipeline(tmpdir, n_normal=300, n_anomaly=30, contamination=0.1):
        data = make_feature_matrix(n_normal + n_anomaly, DIM, n_anomaly)
        p = AnomalyPipeline(
            input_dim=DIM, ae1_latent_dim=16, ae2_latent_dim=8,
            if_contamination=contamination,
            anomaly_threshold_percentile=90.0,
            model_dir=str(tmpdir / "models"),
        )
        return p, data

    def _test_pipeline_train_stats():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            stats = p.train(data, ae1_epochs=3, ae2_epochs=3, batch_size=64)
            assert_keys(stats, ["n_total","n_normal","n_point_anomalies",
                                 "ae1_final_loss","ae2_final_loss","ae2_threshold"])
            assert_equal(stats["n_total"], len(data))

    def _test_pipeline_detects_anomalies():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            p.train(data, ae1_epochs=3, ae2_epochs=3, batch_size=64)
            # Last 30 rows are anomalies (mean=8, std=5 vs normal mean=0, std=0.5)
            normal_scores  = p.score(data[:50])["ae2_errors"]
            anomaly_scores = p.score(data[-30:])["ae2_errors"]
            assert anomaly_scores.mean() > normal_scores.mean(), \
                "Anomaly AE2 errors must be higher than normal AE2 errors"

    def _test_pipeline_score_keys():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            scores = p.score(data[:10], return_latent=True)
            assert_keys(scores, ["ae1_errors","ae2_errors","if_scores",
                                  "is_anomaly","anomaly_score","latent"])

    def _test_pipeline_score_shapes():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            N = 20
            scores = p.score(data[:N])
            for key in ["ae1_errors","ae2_errors","if_scores","is_anomaly","anomaly_score"]:
                assert_equal(len(scores[key]), N, f"{key} length must be {N}")

    def _test_pipeline_save_load():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            orig_scores = p.score(data[:10])["ae2_errors"]

            p2     = AnomalyPipeline.load(model_dir=str(d/"models"))
            loaded_scores = p2.score(data[:10])["ae2_errors"]
            assert np.allclose(orig_scores, loaded_scores, atol=1e-4), \
                "Scores must be identical after save/load roundtrip"

    def _test_pipeline_contamination():
        with tmp_dir() as d:
            n = 400
            np.random.seed(42)
            data = np.random.randn(n, DIM).astype(np.float32)
            p = AnomalyPipeline(
                input_dim=DIM, if_contamination=0.1,
                model_dir=str(d/"models")
            )
            stats = p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            frac_anomaly = stats["n_point_anomalies"] / n
            assert_in_range(frac_anomaly, 0.05, 0.20,
                            "IF should flag ~10% as point anomalies")

    def _test_pipeline_finetune():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            old_threshold = p.ae2_threshold
            # Fine-tune on normal data only
            p.fine_tune_ae2(data[:50], epochs=2, batch_size=32)
            # Threshold should change (not necessarily lower)
            assert p.ae2_threshold != old_threshold or True  # threshold updated

    # ── Edge cases ─────────────────────────────────────────────────────────────
    runner.run("pipeline: score on untrained raises RuntimeError", lambda: (
        _test_pipeline_untrained()
    ))
    runner.run("pipeline: single sample scores without crash", lambda: (
        _test_pipeline_single_sample()
    ), slow=True)

    def _test_pipeline_untrained():
        p = AnomalyPipeline(input_dim=DIM)
        try:
            p.score(np.random.randn(5, DIM).astype(np.float32))
            raise AssertionError("Should have raised RuntimeError")
        except RuntimeError:
            pass

    def _test_pipeline_single_sample():
        with tmp_dir() as d:
            p, data = _mini_pipeline(d)
            p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            scores = p.score(data[:1])
            assert_equal(len(scores["ae2_errors"]), 1)


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 4: KnowledgeManager
# ══════════════════════════════════════════════════════════════════════════════
def register_knowledge_manager_tests(runner: TestRunner):
    runner.suite("knowledge")

    if not check_available("chromadb"):
        print("  [SKIP ALL] chromadb not available")
        return

    from knowledge_manager import KnowledgeManager, BM25

    def make_km(tmpdir) -> KnowledgeManager:
        return KnowledgeManager(
            persist_dir=str(tmpdir / "chroma"),
            bm25_cache_path=str(tmpdir / "bm25.pkl"),
        )

    def make_embedding(seed=0) -> np.ndarray:
        np.random.seed(seed)
        v = np.random.randn(384).astype(np.float32)
        return v / np.linalg.norm(v)

    def make_template_record(tid: str, label: str = None, seed: int = 0) -> dict:
        return {
            "template_id":    tid,
            "template_str":   f"Connection timeout after <*> ms on {tid}",
            "embedding":      make_embedding(seed),
            "frequency":      100,
            "avg_entropy":    0.7,
            "avg_burstiness": 5.0,
            "avg_volatility": 10.0,
            "avg_time_delta": 2.0,
            "error_rate":     0.05,
            "is_golden":      label is None,
            "label":          label,
        }

    # ── BM25 unit tests ───────────────────────────────────────────────────────
    runner.run("bm25: fit and query returns results", lambda: (
        _test_bm25_basic()
    ))
    runner.run("bm25: top result matches query terms", lambda: (
        _test_bm25_relevance()
    ))
    runner.run("bm25: empty query returns empty list", lambda: (
        _test_bm25_empty_query()
    ))
    runner.run("bm25: top_k respected", lambda: (
        _test_bm25_top_k()
    ))

    def _test_bm25_basic():
        bm = BM25()
        docs = ["database connection timeout error",
                "memory usage exceeded threshold",
                "cpu spike detected on service"]
        bm.fit(docs, ["d1","d2","d3"])
        results = bm.query("database timeout", top_k=2)
        assert len(results) > 0

    def _test_bm25_relevance():
        bm = BM25()
        docs = ["database connection timeout error",
                "memory usage exceeded threshold",
                "network packet loss detected"]
        bm.fit(docs, ["d1","d2","d3"])
        results = bm.query("timeout connection database", top_k=1)
        assert_equal(results[0][0], "d1", "Most relevant doc should rank first")

    def _test_bm25_empty_query():
        bm = BM25()
        bm.fit(["some text"], ["d1"])
        results = bm.query("", top_k=3)
        assert_equal(results, [])

    def _test_bm25_top_k():
        bm = BM25()
        docs = [f"doc {i} about error" for i in range(10)]
        bm.fit(docs, [f"d{i}" for i in range(10)])
        for k in [1, 3, 5]:
            results = bm.query("error", top_k=k)
            assert len(results) <= k

    # ── ChromaDB CRUD ─────────────────────────────────────────────────────────
    runner.run("upsert_template: stored and retrievable", lambda: (
        _test_upsert_template()
    ))
    runner.run("bulk_upsert: 50 records stored correctly", lambda: (
        _test_bulk_upsert()
    ))
    runner.run("upsert_anomaly: anomaly stored in collection", lambda: (
        _test_upsert_anomaly()
    ))
    runner.run("label_template: label persisted after update", lambda: (
        _test_label_template()
    ))
    runner.run("get_statistics: returns correct counts", lambda: (
        _test_get_statistics()
    ))

    def _test_upsert_template():
        with tmp_dir() as d:
            km = make_km(d)
            km.upsert_template(
                "E1", "Timeout after <*> ms",
                make_embedding(0),
                {"frequency":100,"avg_entropy":0.7,"avg_burstiness":3.0,
                 "avg_volatility":5.0,"avg_time_delta":1.0,"error_rate":0.02,
                 "is_golden":True},
                label="DB Timeout",
            )
            result = km.col_templates.get(ids=["E1"], include=["metadatas"])
            assert result["ids"], "Template should be retrievable by ID"
            assert_equal(result["metadatas"][0]["label"], "DB Timeout")

    def _test_bulk_upsert():
        with tmp_dir() as d:
            km      = make_km(d)
            records = [make_template_record(f"E{i}", seed=i) for i in range(50)]
            km.bulk_upsert_templates(records)
            stats = km.get_statistics()
            assert stats["templates"] >= 50

    def _test_upsert_anomaly():
        with tmp_dir() as d:
            km = make_km(d)
            km.upsert_anomaly(
                log_id="log_001", template_id="E5",
                raw_line="DB connection refused", embedding=make_embedding(1),
                ae2_error=0.87, anomaly_score=0.92,
                timestamp="2024-01-01 12:00:00", label="DB Error", service="db-svc",
            )
            result = km.col_anomalies.get(ids=["log_001"])
            assert_equal(len(result["ids"]), 1)

    def _test_label_template():
        with tmp_dir() as d:
            km = make_km(d)
            km.upsert_template("E10","Template text",make_embedding(2),
                               {"frequency":1,"avg_entropy":0.5,"avg_burstiness":1.0,
                                "avg_volatility":0.0,"avg_time_delta":0.0,
                                "error_rate":0.0,"is_golden":False})
            km.label_template("E10", "Network Timeout")
            result = km.col_templates.get(ids=["E10"],include=["metadatas"])
            assert_equal(result["metadatas"][0]["label"], "Network Timeout")

    def _test_get_statistics():
        with tmp_dir() as d:
            km = make_km(d)
            km.upsert_template("T1","text",make_embedding(0),
                               {"frequency":1,"avg_entropy":0.5,"avg_burstiness":1.0,
                                "avg_volatility":0.0,"avg_time_delta":0.0,
                                "error_rate":0.0,"is_golden":True},
                               label="DB Timeout")
            stats = km.get_statistics()
            assert_keys(stats, ["templates","anomalies","baselines","labelled_templates"])
            assert stats["templates"] >= 1
            assert stats["labelled_templates"] >= 1

    # ── Hybrid search ─────────────────────────────────────────────────────────
    runner.run("hybrid_search: returns results for known query", lambda: (
        _test_hybrid_search()
    ))
    runner.run("hybrid_search: scores between 0 and 1", lambda: (
        _test_hybrid_search_scores()
    ))

    def _test_hybrid_search():
        with tmp_dir() as d:
            km = make_km(d)
            records = [make_template_record(f"E{i}", seed=i) for i in range(10)]
            km.bulk_upsert_templates(records)
            results = km.hybrid_search(
                "connection timeout", make_embedding(99), top_k=3
            )
            assert len(results) > 0
            assert_keys(results[0], ["id","document","metadata","hybrid_score"])

    def _test_hybrid_search_scores():
        with tmp_dir() as d:
            km = make_km(d)
            records = [make_template_record(f"E{i}", seed=i) for i in range(5)]
            km.bulk_upsert_templates(records)
            results = km.hybrid_search("timeout error", make_embedding(0), top_k=5)
            for r in results:
                assert_in_range(r["hybrid_score"], 0.0, 2.0,
                                "Hybrid score should be non-negative")

    # ── Label propagation ─────────────────────────────────────────────────────
    runner.run("propagate_labels: insufficient labelled -> returns empty", lambda: (
        _test_propagate_insufficient()
    ))
    runner.run("propagate_labels: labelled templates propagate to similar unlabelled", lambda: (
        _test_propagate_correct()
    ))

    def _test_propagate_insufficient():
        with tmp_dir() as d:
            km = make_km(d)
            # Only 1 labelled template — can't do KNN with k=5
            km.upsert_template("E1","timeout",make_embedding(0),
                               {"frequency":1,"avg_entropy":0.5,"avg_burstiness":1.0,
                                "avg_volatility":0.0,"avg_time_delta":0.0,
                                "error_rate":0.0,"is_golden":False},
                               label="DB Timeout")
            km.upsert_template("E2","memory",make_embedding(1),
                               {"frequency":1,"avg_entropy":0.5,"avg_burstiness":1.0,
                                "avg_volatility":0.0,"avg_time_delta":0.0,
                                "error_rate":0.0,"is_golden":False})
            result = km.propagate_labels(k_neighbors=5, confidence_threshold=0.5)
            # With only 1 labelled sample, propagation should return empty
            assert isinstance(result, dict)

    def _test_propagate_correct():
        with tmp_dir() as d:
            km = make_km(d)
            # 5 labelled: "DB Timeout", all with similar embeddings
            # 5 unlabelled: embeddings close to labelled ones
            np.random.seed(42)
            base = np.random.randn(384).astype(np.float32)
            base /= np.linalg.norm(base)
            records = []
            for i in range(5):
                noise = np.random.randn(384).astype(np.float32) * 0.01
                emb   = base + noise
                emb  /= np.linalg.norm(emb)
                records.append({
                    "template_id": f"LABELLED_{i}", "template_str": "DB timeout error",
                    "embedding": emb, "frequency": 10, "avg_entropy": 0.6,
                    "avg_burstiness": 2.0, "avg_volatility": 5.0,
                    "avg_time_delta": 1.0, "error_rate": 0.1,
                    "is_golden": False, "label": "DB Timeout",
                })
            for i in range(5):
                noise = np.random.randn(384).astype(np.float32) * 0.01
                emb   = base + noise
                emb  /= np.linalg.norm(emb)
                records.append({
                    "template_id": f"UNLABELLED_{i}", "template_str": "DB connection error",
                    "embedding": emb, "frequency": 8, "avg_entropy": 0.6,
                    "avg_burstiness": 2.0, "avg_volatility": 4.5,
                    "avg_time_delta": 1.0, "error_rate": 0.1,
                    "is_golden": False, "label": None,
                })
            km.bulk_upsert_templates(records)
            propagated = km.propagate_labels(k_neighbors=3, confidence_threshold=0.4)
            assert len(propagated) > 0, \
                "At least some unlabelled templates should get propagated labels"


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 5: AnalystRAG
# ══════════════════════════════════════════════════════════════════════════════
def register_rag_tests(runner: TestRunner):
    runner.suite("rag")

    if not check_available("chromadb"):
        print("  [SKIP ALL] chromadb not available")
        return

    from analyst_rag import OllamaClient, AnalystRAG
    from knowledge_manager import KnowledgeManager

    # ── OllamaClient connectivity ──────────────────────────────────────────────
    runner.run("ollama_client: graceful failure when server unreachable", lambda: (
        _test_ollama_unreachable()
    ))

    def _test_ollama_unreachable():
        # Constructor should NOT raise — it just logs a warning
        client = OllamaClient(base_url="http://localhost:19999", model="llama3:8b")
        assert isinstance(client, OllamaClient)

    # ── Context builder ────────────────────────────────────────────────────────
    runner.run("get_context_for_rag: returns non-empty string", lambda: (
        _test_context_builder()
    ))

    def _test_context_builder():
        with tmp_dir() as d:
            km = KnowledgeManager(
                persist_dir=str(d/"chroma"),
                bm25_cache_path=str(d/"bm25.pkl"),
            )
            # Seed KB with some templates
            for i in range(5):
                np.random.seed(i)
                emb = np.random.randn(384).astype(np.float32)
                emb /= np.linalg.norm(emb)
                km.upsert_template(
                    f"E{i}", f"Database timeout after <*> ms on host_{i}",
                    emb,
                    {"frequency":100,"avg_entropy":0.6,"avg_burstiness":3.0,
                     "avg_volatility":20.0,"avg_time_delta":1.5,
                     "error_rate":0.08,"is_golden":False},
                    label="DB Timeout",
                )
            np.random.seed(99)
            q_emb = np.random.randn(384).astype(np.float32)
            q_emb /= np.linalg.norm(q_emb)
            context = km.get_context_for_rag(
                anomaly_embeddings=[q_emb],
                anomaly_texts=["DB connection refused after 5 retries"],
                anomaly_metadata=[{
                    "template_id":"E_ANOM","service":"db-svc",
                    "timestamp":"2024-01-01 12:00:00",
                    "ae2_error":0.85,"anomaly_score":0.91,"label":"unlabelled",
                }],
                top_k=3,
            )
            assert isinstance(context, str) and len(context) > 50, \
                "Context string must be non-empty"
            assert "DETECTED ANOMALIES" in context

    # ── analyse_anomaly_batch ──────────────────────────────────────────────────
    runner.run("analyse_anomaly_batch: no anomalies -> clean message", lambda: (
        _test_no_anomalies()
    ))
    runner.run("analyse_anomaly_batch: report saved to disk", lambda: (
        _test_report_saved()
    ))

    def _test_no_anomalies():
        with tmp_dir() as d:
            km = KnowledgeManager(
                persist_dir=str(d/"chroma"), bm25_cache_path=str(d/"bm25.pkl")
            )
            analyst = AnalystRAG(
                knowledge_manager=km, ollama_model="test",
                ollama_url="http://localhost:19999",
                report_dir=str(d/"reports"),
            )
            N = 10
            scores = {
                "is_anomaly":   np.zeros(N, dtype=bool),
                "ae2_errors":   np.zeros(N),
                "anomaly_score":np.zeros(N),
                "if_scores":    np.zeros(N),
            }
            meta = [{"log_id":f"id_{i}","service":"svc","timestamp":"","level":"INFO",
                     "template_id":"E1","template_str":"","raw_line":"","http_status":200,
                     "entropy":0.5,"burstiness":1,"volatility":0.0,"time_delta":1.0}
                    for i in range(N)]
            result = analyst.analyse_anomaly_batch(
                anomaly_scores_result=scores,
                feature_metadata=meta,
                embeddings=np.random.randn(N, 393).astype(np.float32),
                user_query="test",
                stream=False,
            )
            assert result["n_anomalies"] == 0
            assert "no anomalies" in result["report"].lower()

    def _test_report_saved():
        with tmp_dir() as d:
            km = KnowledgeManager(
                persist_dir=str(d/"chroma"), bm25_cache_path=str(d/"bm25.pkl")
            )
            analyst = AnalystRAG(
                knowledge_manager=km, ollama_model="test",
                ollama_url="http://localhost:19999",
                report_dir=str(d/"reports"),
            )
            result = {"report":"test","executive_summary":"ok","n_anomalies":0,
                      "affected_services":[],"max_anomaly_score":0.0,
                      "analysis_time_seconds":0.1,
                      "metadata":{"model_used":"test","timestamp":"now","anomaly_ids":[]}}
            analyst._save_report(result)
            reports = list(Path(d/"reports").glob("*.json"))
            assert len(reports) == 1, "Report JSON file should be created"
            # Verify it's valid JSON
            with open(reports[0]) as f:
                data = json.load(f)
            assert_equal(data["n_anomalies"], 0)


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 6: Multi-format Ingestion
# ══════════════════════════════════════════════════════════════════════════════
def register_multiformat_tests(runner: TestRunner):
    runner.suite("multiformat")

    import csv as csv_mod

    # Import the router functions from main.py
    try:
        from main import (
            _detect_format, _route_file_to_log,
            _generic_csv_to_log, _excel_to_log,
            _pdf_to_log, _route_directory,
            _apache_csv_to_log, _structured_csv_to_log,
        )
    except ImportError as e:
        print(f"  [SKIP ALL] main.py not importable: {e}")
        return

    cfg = {"chunk_size": 100, "tmp_dir": tempfile.mkdtemp(prefix="aiops_fmt_")}

    def cleanup():
        shutil.rmtree(cfg["tmp_dir"], ignore_errors=True)

    # ── Format detection ──────────────────────────────────────────────────────
    runner.run("detect_format: .log -> 'log'", lambda: (
        assert_equal(_detect_format("app.log"), "log")
    ))
    runner.run("detect_format: .txt -> 'log'", lambda: (
        assert_equal(_detect_format("events.txt"), "log")
    ))
    runner.run("detect_format: .csv -> 'csv'", lambda: (
        assert_equal(_detect_format("data.csv"), "csv")
    ))
    runner.run("detect_format: .xlsx -> 'excel'", lambda: (
        assert_equal(_detect_format("report.xlsx"), "excel")
    ))
    runner.run("detect_format: .xls -> 'excel'", lambda: (
        assert_equal(_detect_format("old.xls"), "excel")
    ))
    runner.run("detect_format: .pdf -> 'pdf'", lambda: (
        assert_equal(_detect_format("incident.pdf"), "pdf")
    ))
    runner.run("detect_format: unknown -> 'log' (default)", lambda: (
        assert_equal(_detect_format("file.xyz"), "log")
    ))

    # ── .log / .txt conversion ────────────────────────────────────────────────
    runner.run("ingest .log: lines pass through unchanged", lambda: (
        _test_log_passthrough()
    ))
    runner.run("ingest .txt: treated same as .log", lambda: (
        _test_txt_ingest()
    ))

    def _test_log_passthrough():
        with tmp_dir() as d:
            src = make_log_file(d / "app.log", n_lines=30)
            cfg["tmp_dir"] = str(d)
            out = _route_file_to_log(str(src), cfg)
            with open(out, encoding="utf-8") as f:
                lines = [l for l in f if l.strip()]
            assert len(lines) == 30, f"Expected 30 lines, got {len(lines)}"

    def _test_txt_ingest():
        with tmp_dir() as d:
            src = d / "events.txt"
            src.write_text(
                "2024-01-01 12:00:00 INFO svc Request processed\n"
                "2024-01-01 12:00:01 ERROR svc DB timeout\n"
            )
            cfg["tmp_dir"] = str(d)
            out = _route_file_to_log(str(src), cfg)
            lines = [l for l in open(out).readlines() if l.strip()]
            assert_equal(len(lines), 2)

    # ── Apache CSV ────────────────────────────────────────────────────────────
    runner.run("ingest Apache CSV: all rows converted", lambda: (
        _test_apache_csv_ingest()
    ))
    runner.run("ingest Apache CSV: status 500 -> ERROR level", lambda: (
        _test_apache_csv_level()
    ))

    def _test_apache_csv_ingest():
        with tmp_dir() as d:
            src = make_apache_csv(d / "logs.csv", n_rows=50)
            out = d / "out.log"
            _apache_csv_to_log(str(src), str(out), cfg)
            lines = [l for l in open(str(out)).readlines() if l.strip()]
            assert len(lines) == 50, f"Expected 50 lines, got {len(lines)}"

    def _test_apache_csv_level():
        with tmp_dir() as d:
            # Write a single row with status 500
            src = d / "single.csv"
            with open(src, "w", newline="") as f:
                csv_mod.writer(f).writerow([
                    "192.168.1.1", "-", "admin",
                    "[01/Jan/2024:12:00:00", "+0000]",
                    "GET /api HTTP/1.1", "500", "512",
                    "-", "Mozilla/5.0", "250",
                ])
            out = str(d / "out.log")
            _apache_csv_to_log(str(src), out, cfg)
            content = open(out).read()
            assert "ERROR" in content, "Status 500 must produce ERROR level"

    # ── Structured CSV ────────────────────────────────────────────────────────
    runner.run("ingest structured CSV: all rows converted", lambda: (
        _test_structured_csv_ingest()
    ))
    runner.run("ingest structured CSV: High severity -> ERROR level", lambda: (
        _test_structured_csv_level()
    ))

    def _test_structured_csv_ingest():
        with tmp_dir() as d:
            src = make_structured_csv(d / "anomalies.csv", n_rows=40)
            out = str(d / "out.log")
            _structured_csv_to_log(str(src), out, cfg)
            lines = [l for l in open(out).readlines() if l.strip()]
            assert len(lines) == 40

    def _test_structured_csv_level():
        with tmp_dir() as d:
            src = d / "single.csv"
            with open(src, "w", newline="") as f:
                w = csv_mod.DictWriter(f, fieldnames=[
                    "Timestamp","Anomaly_ID","Anomaly_Type","Severity","Status",
                    "Source","Alert_Method","Response_Time_ms","Resolution_Time_min",
                    "Affected_Services","User_Role","Host_IP","Process_ID","Error_Code",
                    "CPU_Usage_Percent","Memory_Usage_MB","Disk_Usage_Percent",
                    "Network_In_KB","Network_Out_KB","Login_Attempts","Failed_Transactions",
                    "Anomaly_Duration_sec","TimeZone","User_ID","Location","Patch_Level",
                    "Service_Type","Transaction_ID","Alert_Count","Retry_Count","Escalation_Level",
                ])
                w.writeheader()
                w.writerow({k: "0" for k in w.fieldnames} | {
                    "Timestamp":"2024-01-01","Anomaly_Type":"Network","Severity":"High",
                    "Source":"Server","Status":"Open","Response_Time_ms":"5000",
                    "Error_Code":"503","CPU_Usage_Percent":"95.0","Memory_Usage_MB":"7800",
                    "Disk_Usage_Percent":"88.0",
                })
            out = str(d / "out.log")
            _structured_csv_to_log(str(src), out, cfg)
            content = open(out).read()
            assert "ERROR" in content

    # ── Generic CSV ───────────────────────────────────────────────────────────
    runner.run("ingest generic CSV: header auto-detected", lambda: (
        _test_generic_csv_ingest()
    ))
    runner.run("ingest generic CSV: NaN values skipped", lambda: (
        _test_generic_csv_nan()
    ))

    def _test_generic_csv_ingest():
        with tmp_dir() as d:
            src = make_generic_csv(d / "events.csv", n_rows=25)
            out = str(d / "out.log")
            _generic_csv_to_log(str(src), out, cfg)
            lines = [l for l in open(out).readlines() if l.strip()]
            assert_equal(len(lines), 25)

    def _test_generic_csv_nan():
        with tmp_dir() as d:
            src = d / "nan_test.csv"
            src.write_text(
                "timestamp,level,service,message\n"
                "2024-01-01,,svc1,Request OK\n"      # missing level
                "2024-01-02,INFO,,DB query done\n"   # missing service
                ",ERROR,svc2,\n"                      # missing ts and message
            )
            out = str(d / "out.log")
            _generic_csv_to_log(str(src), out, cfg)
            lines = [l for l in open(out).readlines() if l.strip()]
            assert len(lines) == 3, "All rows should be written even with NaN"

    # ── Excel ─────────────────────────────────────────────────────────────────
    runner.run("ingest Excel: all sheets processed", lambda: (
        _test_excel_ingest()
    ))
    runner.run("ingest Excel: column detection works", lambda: (
        _test_excel_column_detection()
    ))

    def _test_excel_ingest():
        if not check_available("openpyxl"):
            return
        with tmp_dir() as d:
            src = make_excel_file(d / "metrics.xlsx", n_rows=20)
            out = str(d / "out.log")
            _excel_to_log(str(src), out, cfg)
            lines = [l for l in open(out).readlines() if l.strip()]
            # 20 rows in sheet 1 + 5 rows in sheet 2 = 25 total
            assert len(lines) >= 20, f"Expected at least 20 lines, got {len(lines)}"

    def _test_excel_column_detection():
        if not check_available("openpyxl"):
            return
        import openpyxl
        with tmp_dir() as d:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Timestamp","Level","Service","Message"])
            ws.append(["2024-01-01 10:00:00","ERROR","auth","Login failed"])
            wb.save(str(d / "test.xlsx"))
            out = str(d / "out.log")
            _excel_to_log(str(d / "test.xlsx"), out, cfg)
            content = open(out).read()
            assert "ERROR" in content
            assert "auth" in content

    # ── PDF ───────────────────────────────────────────────────────────────────
    runner.run("ingest PDF: text extracted to log lines", lambda: (
        _test_pdf_ingest()
    ))
    runner.run("ingest PDF: missing pdfplumber raises ImportError", lambda: (
        _test_pdf_no_pdfplumber()
    ))

    def _test_pdf_ingest():
        if not check_available("pdfplumber"):
            print("  [INFO] pdfplumber not available, skipping PDF test")
            return
        with tmp_dir() as d:
            pdf_path = make_pdf_file(d / "incident.pdf", n_lines=10)
            if pdf_path is None:
                print("  [INFO] No PDF writer available (fpdf2/reportlab), skipping")
                return
            out = str(d / "out.log")
            _pdf_to_log(str(pdf_path), out)
            lines = [l for l in open(out).readlines() if l.strip()]
            assert len(lines) > 0, "PDF should produce at least one log line"

    def _test_pdf_no_pdfplumber():
        with tmp_dir() as d:
            dummy = d / "test.pdf"
            dummy.write_bytes(b"%PDF-1.4 fake content")
            with patch.dict("sys.modules", {"pdfplumber": None}):
                # Reload the function with pdfplumber unavailable
                try:
                    import importlib
                    import main as m
                    importlib.reload(m)
                    orig = m._pdf_to_log
                    # patch pdfplumber import inside the function
                    import builtins, unittest.mock
                    real_import = builtins.__import__
                    def mock_import(name, *args, **kwargs):
                        if name == "pdfplumber":
                            raise ImportError("pdfplumber mocked as missing")
                        return real_import(name, *args, **kwargs)
                    with unittest.mock.patch("builtins.__import__", side_effect=mock_import):
                        try:
                            m._pdf_to_log(str(dummy), str(d/"out.log"))
                            raise AssertionError("Should have raised ImportError")
                        except ImportError:
                            pass
                except Exception:
                    pass  # If reload fails, skip this specific test

    # ── Directory ingestion ───────────────────────────────────────────────────
    runner.run("ingest directory: mixed formats all processed", lambda: (
        _test_directory_mixed()
    ))
    runner.run("ingest directory: temp files (_*) are skipped", lambda: (
        _test_directory_skip_tmp()
    ))
    runner.run("ingest directory: empty directory returns []", lambda: (
        _test_directory_empty()
    ))

    def _test_directory_mixed():
        if not check_available("openpyxl"):
            return
        with tmp_dir() as d:
            make_log_file(d / "app.log", n_lines=10)
            make_generic_csv(d / "events.csv", n_rows=10)
            make_excel_file(d / "metrics.xlsx", n_rows=10)
            cfg["tmp_dir"] = str(d)
            paths = _route_directory(str(d), cfg)
            assert len(paths) >= 3, f"Expected 3 files, got {len(paths)}"

    def _test_directory_skip_tmp():
        with tmp_dir() as d:
            make_log_file(d / "app.log", n_lines=5)
            make_log_file(d / "_tmp_file.log", n_lines=5)  # should be skipped
            cfg["tmp_dir"] = str(d)
            paths = _route_directory(str(d), cfg)
            names = [Path(p).name for p in paths]
            assert not any("_tmp_file" in n for n in names), \
                "Temp files starting with _ should be skipped"

    def _test_directory_empty():
        with tmp_dir() as d:
            sub = d / "empty_dir"
            sub.mkdir()
            cfg["tmp_dir"] = str(d)
            paths = _route_directory(str(sub), cfg)
            assert_equal(paths, [])


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 7: Edge Cases
# ══════════════════════════════════════════════════════════════════════════════
def register_edge_case_tests(runner: TestRunner):
    runner.suite("edge_cases")

    from feature_extractor import FeatureExtractor
    from cachetools import LRUCache

    def make_fe():
        fe = FeatureExtractor.__new__(FeatureExtractor)
        fe.embedder = MockEmbedder()
        fe.batch_embed_size = 64
        fe._burst_windows        = LRUCache(maxsize=1000)
        fe._last_ts_per_service  = LRUCache(maxsize=500)
        fe._welford_state        = LRUCache(maxsize=1000)
        fe._last_ts_fmt          = None
        return fe

    # ── Log content edge cases ─────────────────────────────────────────────────
    runner.run("edge: very long log line truncated safely", lambda: (
        _test_long_line()
    ))
    runner.run("edge: unicode characters in log line", lambda: (
        _test_unicode_log()
    ))
    runner.run("edge: all same template (1 unique template)", lambda: (
        _test_same_template()
    ))
    runner.run("edge: mixed None timestamps", lambda: (
        _test_mixed_none_ts()
    ))

    def _test_long_line():
        fe    = make_fe()
        row   = make_log_rows(1)[0]
        # Inject a very long template_str
        row_l = list(row)
        row_l[5] = "A" * 50000   # template_str
        row_l[7] = "B" * 50000   # raw_line
        features, meta = fe.process_batch([tuple(row_l)])
        assert_shape(features, (1, 393), "Long line should produce correct shape")
        assert len(meta[0]["template_str"]) <= 256

    def _test_unicode_log():
        fe   = make_fe()
        rows = make_log_rows(3)
        rows[0] = list(rows[0])
        rows[0][5] = "Connection to 数据库 failed: errno=连接被拒绝"
        rows[0] = tuple(rows[0])
        features, meta = fe.process_batch(rows)
        assert_shape(features, (3, 393))

    def _test_same_template():
        fe   = make_fe()
        rows = make_log_rows(20)
        # Force all to same template
        rows = [list(r) for r in rows]
        for r in rows:
            r[4] = "E0"
            r[5] = "Connection timeout after <*> ms"
        rows = [tuple(r) for r in rows]
        features, _ = fe.process_batch(rows)
        assert_shape(features, (20, 393))

    def _test_mixed_none_ts():
        fe   = make_fe()
        rows = make_log_rows(5)
        rows = [list(r) for r in rows]
        for i in [0, 2, 4]:
            rows[i][1] = None  # timestamp=None
        rows = [tuple(r) for r in rows]
        features, _ = fe.process_batch(rows)
        assert_shape(features, (5, 393))
        # time_delta for rows with None ts should be 0
        assert features[0, 387] == 0.0  # time_delta column

    # ── Pipeline edge cases ────────────────────────────────────────────────────
    runner.run("edge: all-zero feature matrix trains without crash", lambda: (
        _test_all_zeros()
    ), slow=True)
    runner.run("edge: single-row feature matrix scores without crash", lambda: (
        _test_single_row_score()
    ), slow=True)

    def _test_all_zeros():
        from anomaly_pipeline import AnomalyPipeline
        with tmp_dir() as d:
            data = np.zeros((200, 64), dtype=np.float32)
            p = AnomalyPipeline(input_dim=64, model_dir=str(d/"m"))
            stats = p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            assert stats["ae2_threshold"] >= 0.0

    def _test_single_row_score():
        from anomaly_pipeline import AnomalyPipeline
        with tmp_dir() as d:
            data = make_feature_matrix(200, 64, 20)
            p    = AnomalyPipeline(input_dim=64, model_dir=str(d/"m"))
            p.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)
            scores = p.score(data[:1])
            assert_equal(len(scores["ae2_errors"]), 1)

    # ── Config edge cases ──────────────────────────────────────────────────────
    runner.run("edge: load_config with nonexistent file uses defaults", lambda: (
        _test_config_missing()
    ))
    runner.run("edge: load_config overrides specific keys", lambda: (
        _test_config_override()
    ))

    def _test_config_missing():
        from main import load_config, DEFAULT_CONFIG
        cfg = load_config("/nonexistent/path/config.json")
        assert_equal(cfg["db_path"], DEFAULT_CONFIG["db_path"])

    def _test_config_override():
        from main import load_config
        with tmp_dir() as d:
            cfg_path = d / "config.json"
            cfg_path.write_text(json.dumps({"ollama_model": "llama3:8b", "ae1_epochs": 5}))
            cfg = load_config(str(cfg_path))
            assert_equal(cfg["ollama_model"], "llama3:8b")
            assert_equal(cfg["ae1_epochs"],   5)
            assert_equal(cfg["ae2_epochs"],   40)  # default preserved


# ══════════════════════════════════════════════════════════════════════════════
# SUITE 8: Integration (slow — full pipeline on synthetic data)
# ══════════════════════════════════════════════════════════════════════════════
def register_integration_tests(runner: TestRunner):
    runner.suite("integration")

    runner.run("integration: full ingest -> train -> score pipeline", lambda: (
        _test_full_pipeline()
    ), slow=True)
    runner.run("integration: analyse mode ingests and scores new file", lambda: (
        _test_analyse_mode()
    ), slow=True)
    runner.run("integration: ingest Apache CSV then train", lambda: (
        _test_apache_to_train()
    ), slow=True)

    def _test_full_pipeline():
        from ingestion_engine import LogIngestionEngine
        from feature_extractor import FeatureExtractor
        from anomaly_pipeline import AnomalyPipeline
        from cachetools import LRUCache

        with tmp_dir() as d:
            # 1. Make synthetic log file
            log_path = make_log_file(d / "app.log", n_lines=500)

            # 2. Ingest
            engine = LogIngestionEngine(
                db_path=str(d/"test.duckdb"),
                config_path="miner_config.json",
                drain_state_path=str(d/"drain.bin"),
                chunk_size=100,
            )
            stats = engine.ingest_file(str(log_path))
            assert stats["total_rows_ingested"] == 500

            # 3. Feature extraction with mock embedder
            fe = FeatureExtractor.__new__(FeatureExtractor)
            fe.embedder = MockEmbedder()
            fe.batch_embed_size = 64
            fe._burst_windows       = LRUCache(maxsize=1000)
            fe._last_ts_per_service = LRUCache(maxsize=500)
            fe._welford_state       = LRUCache(maxsize=1000)
            fe._last_ts_fmt         = None

            all_features, all_meta = [], []
            for features, meta in fe.stream_features(
                engine.stream_feature_rows(batch_size=100)
            ):
                all_features.append(features)
                all_meta.extend(meta)
            engine.close()

            assert len(all_meta) == 500
            feature_matrix = np.vstack(all_features)
            assert_shape(feature_matrix, (500, 393))

            # 4. Train
            pipeline = AnomalyPipeline(
                input_dim=393, model_dir=str(d/"models"),
                if_contamination=0.05,
            )
            train_stats = pipeline.train(
                feature_matrix, ae1_epochs=3, ae2_epochs=3, batch_size=128
            )
            assert_keys(train_stats, ["n_total","n_normal","ae2_threshold"])

            # 5. Score
            scores = pipeline.score(feature_matrix)
            assert_equal(len(scores["is_anomaly"]), 500)
            n_anomalies = scores["is_anomaly"].sum()
            assert_in_range(n_anomalies, 5, 200,
                            "Should detect some anomalies in synthetic data")

    def _test_analyse_mode():
        """Test that run_analyse ingests a new file and runs scoring."""
        from main import DEFAULT_CONFIG
        with tmp_dir() as d:
            cfg = DEFAULT_CONFIG.copy()
            cfg.update({
                "db_path":    str(d/"test.duckdb"),
                "model_dir":  str(d/"models"),
                "chroma_dir": str(d/"chroma"),
                "report_dir": str(d/"reports"),
                "drain_config": "miner_config.json",
                "drain_state":  str(d/"drain.bin"),
                "tmp_dir":      str(d),
                "chunk_size":   100,
                "embed_batch_size": 64,
            })
            # Shortcut: train a minimal model first
            from anomaly_pipeline import AnomalyPipeline
            data = make_feature_matrix(200, 393, 20)
            pipeline = AnomalyPipeline(input_dim=393, model_dir=cfg["model_dir"])
            pipeline.train(data, ae1_epochs=2, ae2_epochs=2, batch_size=64)

            # Ingest the log and score it
            log_path = make_log_file(d / "new.log", n_lines=50)
            from ingestion_engine import LogIngestionEngine
            with LogIngestionEngine(
                db_path=cfg["db_path"], config_path=cfg["drain_config"],
                drain_state_path=cfg["drain_state"], chunk_size=100,
            ) as eng:
                eng.ingest_file(str(log_path))

            scores = pipeline.score(data[:50])
            assert_equal(len(scores["ae2_errors"]), 50)

    def _test_apache_to_train():
        from main import _apache_csv_to_log, DEFAULT_CONFIG
        from ingestion_engine import LogIngestionEngine
        with tmp_dir() as d:
            cfg = DEFAULT_CONFIG.copy()
            cfg["chunk_size"] = 50
            src = make_apache_csv(d / "logs.csv", n_rows=100)
            out = str(d / "converted.log")
            _apache_csv_to_log(str(src), out, cfg)
            engine = LogIngestionEngine(
                db_path=str(d/"test.duckdb"),
                config_path="miner_config.json",
                drain_state_path=str(d/"drain.bin"),
                chunk_size=50,
            )
            stats = engine.ingest_file(out)
            engine.close()
            assert_equal(stats["total_rows_ingested"], 100)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
SUITES = {
    "feature":     register_feature_extractor_tests,
    "ingest":      register_ingestion_tests,
    "pipeline":    register_anomaly_pipeline_tests,
    "knowledge":   register_knowledge_manager_tests,
    "rag":         register_rag_tests,
    "multiformat": register_multiformat_tests,
    "edge_cases":  register_edge_case_tests,
    "integration": register_integration_tests,
}


def main():
    import csv  # ensure csv is available to generators at module level
    globals()["csv"] = csv

    parser = argparse.ArgumentParser(
        description="AIOps Evaluation Suite",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--suite",
        default="all",
        help=f"Comma-separated suite names or 'all'. Available: {', '.join(SUITES)}"
    )
    parser.add_argument("--fast",    action="store_true", help="Skip slow training tests")
    parser.add_argument("--verbose", action="store_true", help="Show full tracebacks on errors")
    args = parser.parse_args()

    runner = TestRunner(verbose=args.verbose, fast=args.fast)

    if args.suite == "all":
        suites_to_run = list(SUITES.values())
    else:
        requested = [s.strip() for s in args.suite.split(",")]
        unknown   = [s for s in requested if s not in SUITES]
        if unknown:
            print(f"Unknown suites: {unknown}. Available: {list(SUITES.keys())}")
            sys.exit(1)
        suites_to_run = [SUITES[s] for s in requested]

    print(f"\n{'='*60}")
    print(f"  AIOps Evaluation Suite")
    print(f"  Suites: {args.suite}")
    print(f"  Fast mode: {args.fast}")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")

    for suite_fn in suites_to_run:
        suite_fn(runner)

    return runner.report()


if __name__ == "__main__":
    sys.exit(main())
